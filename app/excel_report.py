"""Excel report generation utilities."""

from __future__ import annotations

import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.reader import get_summary_stats


def _pick_chart_columns(df: pd.DataFrame) -> tuple[str | None, str | None]:
    numeric_columns = list(df.select_dtypes(include="number").columns)
    if not numeric_columns:
        return None, None
    y_col = str(numeric_columns[0])

    x_col = None
    for col in df.columns:
        if str(col) == y_col:
            continue
        if pd.api.types.is_object_dtype(df[col]) or pd.api.types.is_datetime64_any_dtype(df[col]):
            x_col = str(col)
            break
    if x_col is None:
        x_col = str(df.columns[0])

    return x_col, y_col


def generate_excel_report(
    df: pd.DataFrame,
    title: str = "Data Report",
    company_name: str = "Report Generator",
    output_path: str = None,
) -> str:
    """
    Generate a styled Excel report workbook from a DataFrame and save it to disk.
    """
    os.makedirs("outputs", exist_ok=True)
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        output_path = os.path.join("outputs", f"report_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    max_col = max(len(df.columns), 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws["A1"] = f"{company_name} — {title}"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(fill_type="solid", start_color="1A2744", end_color="1A2744")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws["A2"] = (
        f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | "
        f"Rows: {len(df)} | Cols: {len(df.columns)}"
    )
    ws["A2"].font = Font(italic=True, size=10, color="444444")
    ws["A2"].fill = PatternFill(fill_type="solid", start_color="F5F5F5", end_color="F5F5F5")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    header_row = 4
    thin_border = Border(
        left=Side(style="thin", color="BDBDBD"),
        right=Side(style="thin", color="BDBDBD"),
        top=Side(style="thin", color="BDBDBD"),
        bottom=Side(style="thin", color="BDBDBD"),
    )

    for col_idx, column_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=str(column_name))
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", start_color="1A2744", end_color="1A2744")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[header_row].height = 20

    for row_offset, row_values in enumerate(df.itertuples(index=False), start=1):
        row_number = header_row + row_offset
        fill_color = "EEF2FF" if row_offset % 2 == 0 else "FFFFFF"
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_number, column=col_idx, value=value)
            cell.fill = PatternFill(fill_type="solid", start_color=fill_color, end_color=fill_color)
            cell.border = thin_border

            if pd.notna(value) and isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0.00" if isinstance(value, float) else "#,##0"
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for col_idx, column_name in enumerate(df.columns, start=1):
        column_data = df[column_name].astype(str).fillna("")
        max_data_len = int(column_data.map(len).max()) if not column_data.empty else 0
        width = min(max(len(str(column_name)), max_data_len) + 4, 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    summary_ws = wb.create_sheet("Summary Stats")
    summary_ws.append(["Metric", "Value"])
    summary_ws["A1"].font = Font(bold=True, color="FFFFFF")
    summary_ws["B1"].font = Font(bold=True, color="FFFFFF")
    summary_ws["A1"].fill = PatternFill(fill_type="solid", start_color="1A2744", end_color="1A2744")
    summary_ws["B1"].fill = PatternFill(fill_type="solid", start_color="1A2744", end_color="1A2744")

    stats = get_summary_stats(df)
    summary_rows = [
        ["Total Records", stats["row_count"]],
        ["Total Columns", stats["col_count"]],
        ["Numeric Columns Count", len(stats["numeric_summary"])],
    ]
    for col, col_stats in stats["numeric_summary"].items():
        summary_rows.extend(
            [
                [f"{col} Min", col_stats["min"]],
                [f"{col} Max", col_stats["max"]],
                [f"{col} Mean", col_stats["mean"]],
                [f"{col} Sum", col_stats["sum"]],
            ]
        )
    for row in summary_rows:
        summary_ws.append(row)

    for row in summary_ws.iter_rows(min_row=1, max_row=summary_ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "#,##0.00"
                else:
                    cell.alignment = Alignment(horizontal="left")
    summary_ws.column_dimensions["A"].width = 32
    summary_ws.column_dimensions["B"].width = 20

    chart_ws = wb.create_sheet("Chart")
    chart_ws.append([str(df.columns[0]), str(df.columns[1]) if len(df.columns) > 1 else "Value"])
    x_col, y_col = _pick_chart_columns(df)
    if x_col and y_col:
        chart_df = df[[x_col, y_col]].dropna().head(30)
        chart_ws.delete_rows(1, 1)
        chart_ws.append([x_col, y_col])
        for row in chart_df.itertuples(index=False):
            chart_ws.append(list(row))

        bar = BarChart()
        bar.title = title
        bar.style = 10
        bar.y_axis.title = y_col
        bar.x_axis.title = x_col
        data_ref = Reference(chart_ws, min_col=2, min_row=1, max_row=chart_ws.max_row)
        categories_ref = Reference(chart_ws, min_col=1, min_row=2, max_row=chart_ws.max_row)
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(categories_ref)
        bar.width = 18
        bar.height = 9
        if bar.series:
            bar.series[0].graphicalProperties.solidFill = "1A2744"
            bar.series[0].graphicalProperties.line.solidFill = "1A2744"
        chart_ws.add_chart(bar, "D2")

    wb.save(output_path)
    return output_path
